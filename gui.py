import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import sys
import json
from pathlib import Path
import numpy as np
import pandas as pd
from scipy.optimize import minimize
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLineEdit,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QMessageBox,
    QLabel,
    QSpinBox,
    QStackedWidget,
    QInputDialog,
    QSizePolicy,
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QCursor, QColor, QBrush, QPen, QFont
import pyqtgraph as pg

PORTFOLIO_DIR = Path.home() / ".efficient_frontier" / "portfolios"

pg.setConfigOption('background', 'k')
pg.setConfigOption('foreground', 'w')

from data_fetch import get_ticker_data


DEFAULT_TICKERS = [
    "AAPL",   # Apple - Technology
    "MSFT",   # Microsoft - Technology
    "GOOGL",  # Alphabet - Technology
    "AMZN",   # Amazon - Consumer Discretionary
    "NVDA",   # NVIDIA - Technology
    "META",   # Meta - Technology
    "BRK-B",  # Berkshire Hathaway - Financials
    "JPM",    # JPMorgan Chase - Financials
    "JNJ",    # Johnson & Johnson - Healthcare
    "V",      # Visa - Financials
    "PG",     # Procter & Gamble - Consumer Staples
    "UNH",    # UnitedHealth - Healthcare
    "HD",     # Home Depot - Consumer Discretionary
    "MA",     # Mastercard - Financials
    "XOM",    # Exxon Mobil - Energy
    "PFE",    # Pfizer - Healthcare
    "KO",     # Coca-Cola - Consumer Staples
    "PEP",    # PepsiCo - Consumer Staples
    "DIS",    # Disney - Communication Services
    "NFLX",   # Netflix - Communication Services
]


class PercentAxisItem(pg.AxisItem):
    """Axis item that formats values as percentages (e.g. 0.15 → 15%)."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.enableAutoSIPrefix(False)

    def tickStrings(self, values, scale, spacing):
        return [f"{v * 100:.0f}%" for v in values]


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Efficient Frontier")
        self.setMinimumSize(900, 600)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        main_layout = QHBoxLayout(central_widget)

        # Left panel for controls
        self.left_panel = QWidget()
        left_layout = QVBoxLayout(self.left_panel)
        self.left_panel.setMaximumWidth(300)

        # Navigation buttons
        nav_layout = QHBoxLayout()
        nav_layout.setSpacing(2)
        nav_layout.setContentsMargins(0, 0, 0, 0)
        self.frontier_button = QPushButton("Frontier")
        self.frontier_button.clicked.connect(lambda: self.show_page("frontier"))
        self.corr_button = QPushButton("Correlation")
        self.corr_button.clicked.connect(self.show_correlation_matrix)
        self.cov_button = QPushButton("Covariance")
        self.cov_button.clicked.connect(self.show_covariance_matrix)
        for btn in (self.frontier_button, self.corr_button, self.cov_button):
            btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        nav_layout.addWidget(self.frontier_button)
        nav_layout.addWidget(self.corr_button)
        nav_layout.addWidget(self.cov_button)
        left_layout.addLayout(nav_layout)

        # Portfolio buttons
        portfolio_layout = QHBoxLayout()
        portfolio_layout.setSpacing(2)
        portfolio_layout.setContentsMargins(0, 0, 0, 0)
        load_button = QPushButton("Load")
        load_button.clicked.connect(self.load_portfolio)
        save_button = QPushButton("Save")
        save_button.clicked.connect(self.save_portfolio)
        clear_button = QPushButton("Clear")
        clear_button.clicked.connect(self.clear_portfolio)
        for btn in (load_button, save_button, clear_button):
            btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        portfolio_layout.addWidget(load_button)
        portfolio_layout.addWidget(save_button)
        portfolio_layout.addWidget(clear_button)
        left_layout.addLayout(portfolio_layout)

        # Input row
        input_layout = QHBoxLayout()
        self.ticker_input = QLineEdit()
        self.ticker_input.setPlaceholderText("Enter ticker (e.g., AAPL)")
        self.ticker_input.returnPressed.connect(self.add_security)

        add_button = QPushButton("Add")
        add_button.clicked.connect(self.add_security)

        input_layout.addWidget(self.ticker_input)
        input_layout.addWidget(add_button)
        left_layout.addLayout(input_layout)

        # Ticker table
        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Ticker", ""])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        left_layout.addWidget(self.table)

        # Simulations input
        sim_layout = QHBoxLayout()
        sim_label = QLabel("Simulations:")
        self.sim_spinbox = QSpinBox()
        self.sim_spinbox.setRange(100, 100000)
        self.sim_spinbox.setValue(50000)
        self.sim_spinbox.setSingleStep(1000)
        sim_layout.addWidget(sim_label)
        sim_layout.addWidget(self.sim_spinbox)
        left_layout.addLayout(sim_layout)

        # Lookback period selector
        lookback_layout = QHBoxLayout()
        lookback_label = QLabel("Lookback:")
        lookback_layout.addWidget(lookback_label)

        self.lookback_buttons = {}
        LOOKBACK_OPTIONS = [("1y", 365), ("2y", 730), ("5y", 1825), ("Max", None)]
        for label, days in LOOKBACK_OPTIONS:
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.clicked.connect(lambda checked, l=label: self.set_lookback(l))
            lookback_layout.addWidget(btn)
            self.lookback_buttons[label] = btn

        self.lookback_buttons["5y"].setChecked(True)
        self.selected_lookback = 1825
        left_layout.addLayout(lookback_layout)

        # Run button
        self.run_button = QPushButton("Run")
        self.run_button.clicked.connect(self.run_optimization)
        self.run_button.setMinimumHeight(40)
        left_layout.addWidget(self.run_button)

        main_layout.addWidget(self.left_panel)

        # Left toggle button
        self.left_toggle = QPushButton("◀")
        self.left_toggle.setFixedWidth(16)
        self.left_toggle.setStyleSheet("QPushButton { border: none; color: #888; font-size: 14px; }")
        self.left_toggle.clicked.connect(self.toggle_left_panel)
        main_layout.addWidget(self.left_toggle)

        # Stacked widget for page switching
        self.stack = QStackedWidget()
        main_layout.addWidget(self.stack, stretch=1)

        # --- Page 0: Frontier (plot + toggle + weights table) ---
        frontier_page = QWidget()
        frontier_layout = QHBoxLayout(frontier_page)
        frontier_layout.setContentsMargins(0, 0, 0, 0)

        self.plot_widget = pg.PlotWidget(
            axisItems={
                'bottom': PercentAxisItem(orientation='bottom'),
                'left': PercentAxisItem(orientation='left'),
            }
        )
        self.plot_widget.setLabel('bottom', 'Standard Deviation')
        self.plot_widget.setLabel('left', 'CAGR')
        self.plot_widget.setTitle('Efficient Frontier')
        self.plot_widget.showGrid(x=False, y=False)
        frontier_layout.addWidget(self.plot_widget, stretch=1)

        self.right_toggle = QPushButton("▶")
        self.right_toggle.setFixedWidth(16)
        self.right_toggle.setStyleSheet("QPushButton { border: none; color: #888; font-size: 14px; }")
        self.right_toggle.clicked.connect(self.toggle_right_panel)
        self.right_toggle.hide()
        frontier_layout.addWidget(self.right_toggle)

        self.weights_table = QTableWidget()
        self.weights_table.setColumnCount(2)  # was 3 with Max Sortino
        self.weights_table.setHorizontalHeaderLabels(["Max Sharpe", "Min Variance"])  # "Max Sortino" hidden temporarily
        self.weights_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.weights_table.verticalHeader().setVisible(True)
        self.weights_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.weights_table.setMaximumWidth(380)
        self.weights_table.setMinimumWidth(340)
        self.weights_table.hide()
        frontier_layout.addWidget(self.weights_table)

        self.stack.addWidget(frontier_page)  # index 0

        # --- Page 1: Correlation Matrix ---
        corr_page = QWidget()
        corr_layout = QVBoxLayout(corr_page)
        corr_title = QLabel("Correlation Matrix")
        corr_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        font = corr_title.font()
        font.setBold(True)
        corr_title.setFont(font)
        corr_layout.addWidget(corr_title)
        self.corr_plot = pg.PlotWidget()
        self.corr_plot.hideButtons()
        self.corr_image = pg.ImageItem()
        self.corr_plot.addItem(self.corr_image)
        self.corr_text_items = []
        corr_layout.addWidget(self.corr_plot)
        self.stack.addWidget(corr_page)  # index 1

        # --- Page 2: Covariance Matrix ---
        cov_page = QWidget()
        cov_layout = QVBoxLayout(cov_page)
        cov_title = QLabel("Covariance Matrix")
        cov_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        font = cov_title.font()
        font.setBold(True)
        cov_title.setFont(font)
        cov_layout.addWidget(cov_title)
        self.cov_plot = pg.PlotWidget()
        self.cov_plot.hideButtons()
        self.cov_image = pg.ImageItem()
        self.cov_plot.addItem(self.cov_image)
        self.cov_text_items = []
        cov_layout.addWidget(self.cov_plot)
        self.stack.addWidget(cov_page)  # index 2

        self.loaded_portfolio_name = None

        self.update_lookback_styles()

    LOOKBACK_MAP = {"1y": 365, "2y": 730, "5y": 1825, "Max": None}

    def set_lookback(self, label):
        self.selected_lookback = self.LOOKBACK_MAP[label]
        for key, btn in self.lookback_buttons.items():
            btn.setChecked(key == label)
        self.update_lookback_styles()

    def update_lookback_styles(self):
        for key, btn in self.lookback_buttons.items():
            if btn.isChecked():
                btn.setStyleSheet("QPushButton { background-color: #1a6ea0; color: white; }")
            else:
                btn.setStyleSheet("")

    def show_page(self, page):
        pages = {"frontier": 0, "correlation": 1, "covariance": 2}
        self.stack.setCurrentIndex(pages[page])

    def toggle_left_panel(self):
        if self.left_panel.isVisible():
            self.left_panel.hide()
            self.left_toggle.setText("▶")
        else:
            self.left_panel.show()
            self.left_toggle.setText("◀")

    def toggle_right_panel(self):
        if self.weights_table.isVisible():
            self.weights_table.hide()
            self.right_toggle.setText("◀")
        else:
            self.weights_table.show()
            self.right_toggle.setText("▶")

    def add_ticker_to_table(self, ticker):
        """Add a ticker to the table without validation."""
        row = self.table.rowCount()
        self.table.insertRow(row)

        ticker_item = QTableWidgetItem(ticker)
        ticker_item.setFlags(ticker_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.table.setItem(row, 0, ticker_item)

        delete_button = QPushButton("Delete")
        delete_button.clicked.connect(lambda checked, r=row: self.remove_security(r))
        self.table.setCellWidget(row, 1, delete_button)

    def add_security(self):
        ticker = self.ticker_input.text().strip().upper()

        if not ticker:
            QMessageBox.warning(self, "Invalid Input", "Please enter a ticker symbol.")
            return

        # Check for duplicates
        for row in range(self.table.rowCount()):
            existing = self.table.item(row, 0)
            if existing and existing.text() == ticker:
                QMessageBox.warning(self, "Duplicate", f"{ticker} is already in the list.")
                self.ticker_input.clear()
                return

        # Validate ticker by fetching data
        QApplication.setOverrideCursor(QCursor(Qt.CursorShape.WaitCursor))
        try:
            data = get_ticker_data(ticker)
            if data.empty:
                QApplication.restoreOverrideCursor()
                QMessageBox.warning(
                    self,
                    "Invalid Ticker",
                    f"No data found for '{ticker}'. Please check the ticker symbol.",
                )
                return
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(
                self,
                "Error",
                f"Failed to fetch data for '{ticker}':\n{str(e)}",
            )
            return
        finally:
            QApplication.restoreOverrideCursor()

        row = self.table.rowCount()
        self.table.insertRow(row)

        ticker_item = QTableWidgetItem(ticker)
        ticker_item.setFlags(ticker_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.table.setItem(row, 0, ticker_item)

        delete_button = QPushButton("Delete")
        delete_button.clicked.connect(lambda checked, r=row: self.remove_security(r))
        self.table.setCellWidget(row, 1, delete_button)

        self.ticker_input.clear()
        self.ticker_input.setFocus()

    def remove_security(self, row):
        sender = self.sender()
        for r in range(self.table.rowCount()):
            if self.table.cellWidget(r, 1) is sender:
                self.table.removeRow(r)
                return

    def get_tickers(self):
        tickers = []
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 0)
            if item:
                tickers.append(item.text())
        return tickers

    def clear_portfolio(self):
        self.table.setRowCount(0)
        self.loaded_portfolio_name = None

    def save_portfolio(self):
        tickers = self.get_tickers()
        if not tickers:
            QMessageBox.warning(self, "Empty Portfolio", "There are no tickers to save.")
            return

        name, ok = QInputDialog.getText(self, "Save Portfolio", "Portfolio name:")
        if not ok or not name.strip():
            return
        name = name.strip()

        try:
            PORTFOLIO_DIR.mkdir(parents=True, exist_ok=True)
            path = PORTFOLIO_DIR / f"{name}.json"
            path.write_text(json.dumps(tickers, indent=2))
            pass  # silently succeed
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save portfolio:\n{e}")

    def load_portfolio(self):
        if not PORTFOLIO_DIR.exists():
            QMessageBox.warning(self, "No Portfolios", "No saved portfolios found.")
            return

        files = sorted(PORTFOLIO_DIR.glob("*.json"))
        if not files:
            QMessageBox.warning(self, "No Portfolios", "No saved portfolios found.")
            return

        names = [f.stem for f in files]
        name, ok = QInputDialog.getItem(self, "Load Portfolio", "Select portfolio:", names, 0, False)
        if not ok:
            return

        try:
            path = PORTFOLIO_DIR / f"{name}.json"
            tickers = json.loads(path.read_text())
            self.clear_portfolio()
            for t in tickers:
                self.add_ticker_to_table(t)
            self.loaded_portfolio_name = name
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load portfolio:\n{e}")

    @staticmethod
    def _sanitize_name(name):
        """Replace spaces and special characters with underscores."""
        return re.sub(r'[^\w]+', '_', name).strip('_')

    def export_outputs(self, tickers):
        """Export xlsx and PNGs to output/<sanitized_portfolio_name>/."""
        if self.loaded_portfolio_name is None:
            return

        sanitized = self._sanitize_name(self.loaded_portfolio_name)
        out_dir = Path("output") / sanitized
        out_dir.mkdir(parents=True, exist_ok=True)

        # Export daily returns xlsx
        try:
            _, daily_returns = self.compute_daily_returns(tickers, self.selected_lookback)
            daily_returns.index.name = "Date"
            daily_returns.sort_index(ascending=False, inplace=True)

            xlsx_path = out_dir / f"{sanitized}.xlsx"
            daily_returns.to_excel(xlsx_path)

            # Add Excel Table object over the written data
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            max_row = ws.max_row
            max_col = ws.max_column
            table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
            tbl = Table(displayName=sanitized, ref=table_ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tbl)
            wb.save(xlsx_path)
        except Exception:
            pass

        # Export PNGs by grabbing each widget
        for widget, suffix in [
            (self.plot_widget, "_Frontier"),
            (self.corr_plot, "_Correlation"),
            (self.cov_plot, "_Covariance"),
        ]:
            try:
                widget.grab().save(str(out_dir / f"{sanitized}{suffix}.png"))
            except Exception:
                pass

    def get_risk_free_rate(self):
        """Get risk-free rate from ^IRX (13-week Treasury Bill)."""
        try:
            data = get_ticker_data("^IRX")
            if data.empty:
                return 0.05  # Default 5% if unavailable
            # ^IRX is quoted as a percentage, convert to decimal
            latest_rate = data["Close"].iloc[-1] / 100
            # Handle multi-index columns from yfinance
            if hasattr(latest_rate, "iloc"):
                latest_rate = latest_rate.iloc[0]
            return float(latest_rate)
        except Exception:
            return 0.05  # Default fallback

    def populate_matrix_plot(self, plot_widget, image_item, text_items_attr, matrix, value_format):
        """Render a DataFrame matrix as a pyqtgraph heatmap with text overlays."""
        labels = list(matrix.index)
        n = len(labels)
        vals = matrix.values

        # Remove old text items
        old_items = getattr(self, text_items_attr)
        for item in old_items:
            plot_widget.removeItem(item)
        setattr(self, text_items_attr, [])

        # Normalize using only lower triangle (exclude diagonal)
        lower = [vals[i, j] for i in range(n) for j in range(i)]
        if lower:
            v_min, v_max = min(lower), max(lower)
        else:
            v_min, v_max = float(vals.min()), float(vals.max())
        v_range = v_max - v_min if v_max != v_min else 1.0

        # Build RGBA image array (N x N x 4)
        # pyqtgraph ImageItem: row=x, col=y, so we build as [col][row] = [x][y]
        # with Y inverted, image[x][y] maps to cell (row=y, col=x)
        img = np.zeros((n, n, 4), dtype=np.uint8)

        # Green-yellow-red colormap matching the old scheme
        cmap_positions = [0.0, 0.5, 1.0]
        cmap_colors = [
            (0, 128, 0, 255),      # green (low)
            (255, 255, 0, 255),    # yellow (mid)
            (215, 48, 39, 255),    # red (high)
        ]
        colormap = pg.ColorMap(cmap_positions, cmap_colors)

        new_text_items = []
        label_font = QFont()
        label_font.setPointSize(10)

        for i in range(n):      # row
            for j in range(n):  # col
                if j >= i:
                    # Upper triangle + diagonal — transparent
                    img[j, i] = [0, 0, 0, 0]
                else:
                    val = vals[i, j]
                    norm = max(0.0, min(1.0, (val - v_min) / v_range))
                    rgba = colormap.map([norm], mode='byte')[0]
                    img[j, i] = rgba

                    # Determine text color based on luminance
                    r, g, b = int(rgba[0]), int(rgba[1]), int(rgba[2])
                    luminance = 0.299 * r + 0.587 * g + 0.114 * b
                    text_color = (0, 0, 0) if luminance > 128 else (255, 255, 255)

                    # Add text overlay centered in cell
                    text = pg.TextItem(
                        text=f"{val:{value_format}}",
                        color=text_color,
                        anchor=(0.5, 0.5),
                    )
                    text.setFont(label_font)
                    text.setPos(j + 0.5, i + 0.5)
                    plot_widget.addItem(text)
                    new_text_items.append(text)

        setattr(self, text_items_attr, new_text_items)

        image_item.setImage(img)
        image_item.setRect(0, 0, n, n)

        # Configure axes
        bottom_ax = plot_widget.getAxis('bottom')
        left_ax = plot_widget.getAxis('left')
        bottom_ticks = [(j + 0.5, labels[j]) for j in range(n)]
        left_ticks = [(i + 0.5, labels[i]) for i in range(n)]
        bottom_ax.setTicks([bottom_ticks])
        left_ax.setTicks([left_ticks])
        bottom_ax.setStyle(tickLength=0, stopAxisAtTick=(True, True))
        left_ax.setStyle(tickLength=0, stopAxisAtTick=(True, True))

        # Invert Y so row 0 is at the top
        plot_widget.getViewBox().invertY(True)
        plot_widget.setXRange(0, n, padding=0)
        plot_widget.setYRange(0, n, padding=0)

    def show_correlation_matrix(self):
        """Compute and display the correlation matrix heatmap."""
        tickers = self.get_tickers()
        if len(tickers) < 2:
            QMessageBox.warning(self, "Insufficient Securities",
                                "Please add at least 2 securities.")
            return

        QApplication.setOverrideCursor(QCursor(Qt.CursorShape.WaitCursor))
        try:
            _, daily_returns = self.compute_daily_returns(tickers, self.selected_lookback)
            corr = daily_returns.corr()
            self.populate_matrix_plot(self.corr_plot, self.corr_image, 'corr_text_items', corr, ".2f")
            self.left_panel.hide()
            self.left_toggle.setText("▶")
            self.show_page("correlation")
        except Exception as e:
            QMessageBox.critical(self, "Error",
                                 f"Failed to compute correlation matrix:\n{str(e)}")
        finally:
            QApplication.restoreOverrideCursor()

    def show_covariance_matrix(self):
        """Compute and display the annualized covariance matrix heatmap."""
        tickers = self.get_tickers()
        if len(tickers) < 2:
            QMessageBox.warning(self, "Insufficient Securities",
                                "Please add at least 2 securities.")
            return

        QApplication.setOverrideCursor(QCursor(Qt.CursorShape.WaitCursor))
        try:
            _, daily_returns = self.compute_daily_returns(tickers, self.selected_lookback)
            cov = daily_returns.cov() * 252
            self.populate_matrix_plot(self.cov_plot, self.cov_image, 'cov_text_items', cov, ".4f")
            self.left_panel.hide()
            self.left_toggle.setText("▶")
            self.show_page("covariance")
        except Exception as e:
            QMessageBox.critical(self, "Error",
                                 f"Failed to compute covariance matrix:\n{str(e)}")
        finally:
            QApplication.restoreOverrideCursor()

    def compute_daily_returns(self, tickers, lookback_days=1825):
        """Fetch price data and compute daily returns for given tickers."""
        price_data = {}
        for ticker in tickers:
            data = get_ticker_data(ticker)
            if isinstance(data.columns, pd.MultiIndex):
                close = data["Close"][ticker]
            else:
                close = data["Close"]
            price_data[ticker] = close

        prices = pd.DataFrame(price_data)
        if lookback_days is not None:
            cutoff = prices.index.max() - pd.Timedelta(days=lookback_days)
            prices = prices.loc[prices.index >= cutoff]
        prices = prices.dropna()

        daily_returns = prices.pct_change().dropna()
        return prices, daily_returns

    def calculate_efficient_frontier(self, tickers, num_simulations, lookback_days=1825):
        """Calculate the efficient frontier for given tickers using Monte Carlo simulation."""
        prices, daily_returns = self.compute_daily_returns(tickers, lookback_days)

        # Calculate CAGR for each asset (more accurate than mean * 252)
        num_years = len(prices) / 252
        cagr = (prices.iloc[-1] / prices.iloc[0]) ** (1 / num_years) - 1
        mean_returns = cagr

        # Annualized covariance matrix
        cov_matrix = daily_returns.cov() * 252

        num_assets = len(tickers)

        def portfolio_volatility(weights):
            return np.sqrt(np.dot(weights.T, np.dot(cov_matrix.values, weights)))

        def portfolio_return(weights):
            return np.dot(weights, mean_returns.values)

        def neg_sharpe_ratio(weights, risk_free_rate):
            ret = portfolio_return(weights)
            vol = portfolio_volatility(weights)
            return -(ret - risk_free_rate) / vol

        # Monte Carlo simulation
        sim_returns = []
        sim_volatilities = []
        sim_sharpe_ratios = []
        risk_free_rate = self.get_risk_free_rate()

        # Pre-generate varying Dirichlet concentrations (log-uniform from 0.01 to 5)
        # Low alpha → concentrated portfolios near individual assets
        # High alpha → diversified portfolios near the center
        alphas = 10 ** np.random.uniform(-2, 0.7, size=num_simulations)

        for alpha in alphas:
            weights = np.random.dirichlet(np.ones(num_assets) * alpha)

            ret = portfolio_return(weights)
            vol = portfolio_volatility(weights)
            sharpe = (ret - risk_free_rate) / vol

            sim_returns.append(ret)
            sim_volatilities.append(vol)
            sim_sharpe_ratios.append(sharpe)

        # Optimization for efficient frontier curve
        constraints = {"type": "eq", "fun": lambda x: np.sum(x) - 1}
        bounds = tuple((0, 1) for _ in range(num_assets))
        initial_weights = np.array([1 / num_assets] * num_assets)

        # Find portfolio with minimum volatility
        min_vol_result = minimize(
            portfolio_volatility,
            initial_weights,
            method="SLSQP",
            bounds=bounds,
            constraints=constraints,
        )
        min_vol_ret = portfolio_return(min_vol_result.x)

        # Find portfolio with maximum return
        max_ret = mean_returns.max()

        # Generate efficient frontier points
        target_returns = np.linspace(min_vol_ret, max_ret, 50)
        frontier_volatilities = []
        frontier_returns = []

        for target in target_returns:
            cons = [
                {"type": "eq", "fun": lambda x: np.sum(x) - 1},
                {"type": "eq", "fun": lambda x, t=target: portfolio_return(x) - t},
            ]
            result = minimize(
                portfolio_volatility,
                initial_weights,
                method="SLSQP",
                bounds=bounds,
                constraints=cons,
            )
            if result.success:
                frontier_volatilities.append(portfolio_volatility(result.x))
                frontier_returns.append(target)

        # Find tangency portfolio (max Sharpe ratio)
        tangency_result = minimize(
            neg_sharpe_ratio,
            initial_weights,
            args=(risk_free_rate,),
            method="SLSQP",
            bounds=bounds,
            constraints=constraints,
        )
        tangency_vol = portfolio_volatility(tangency_result.x)
        tangency_ret = portfolio_return(tangency_result.x)
        tangency_weights = tangency_result.x
        sharpe_ratio = (tangency_ret - risk_free_rate) / tangency_vol

        # Find max Sortino ratio portfolio
        def neg_sortino_ratio(weights):
            port_daily = daily_returns.values @ weights
            downside = port_daily[port_daily < risk_free_rate / 252] - risk_free_rate / 252
            downside_dev = np.sqrt(np.mean(downside ** 2)) * np.sqrt(252)
            if downside_dev < 1e-10:
                return 0.0
            ret = portfolio_return(weights)
            return -(ret - risk_free_rate) / downside_dev

        sortino_result = minimize(
            neg_sortino_ratio,
            initial_weights,
            method="SLSQP",
            bounds=bounds,
            constraints=constraints,
        )
        sortino_weights = sortino_result.x
        sortino_vol = portfolio_volatility(sortino_weights)
        sortino_ret = portfolio_return(sortino_weights)
        sortino_ratio = -neg_sortino_ratio(sortino_weights)

        # Min volatility portfolio stats
        min_vol_weights = min_vol_result.x
        min_vol_vol = portfolio_volatility(min_vol_weights)

        # Individual asset points
        individual_vols = [np.sqrt(cov_matrix.iloc[i, i]) for i in range(num_assets)]
        individual_rets = mean_returns.values.tolist()

        return {
            "frontier_vols": frontier_volatilities,
            "frontier_rets": frontier_returns,
            "sim_vols": sim_volatilities,
            "sim_rets": sim_returns,
            "sim_sharpes": sim_sharpe_ratios,
            "tangency_vol": tangency_vol,
            "tangency_ret": tangency_ret,
            "tangency_weights": tangency_weights,
            "sharpe_ratio": sharpe_ratio,
            "sortino_weights": sortino_weights,
            "sortino_vol": sortino_vol,
            "sortino_ret": sortino_ret,
            "sortino_ratio": sortino_ratio,
            "min_vol_weights": min_vol_weights,
            "min_vol_vol": min_vol_vol,
            "min_vol_ret": min_vol_ret,
            "risk_free_rate": risk_free_rate,
            "individual_vols": individual_vols,
            "individual_rets": individual_rets,
            "tickers": tickers,
        }

    def run_optimization(self):
        tickers = self.get_tickers()

        if len(tickers) < 2:
            QMessageBox.warning(
                self,
                "Insufficient Securities",
                "Please add at least 2 securities to calculate the efficient frontier.",
            )
            return

        num_simulations = self.sim_spinbox.value()

        QApplication.setOverrideCursor(QCursor(Qt.CursorShape.WaitCursor))
        try:
            results = self.calculate_efficient_frontier(tickers, num_simulations, self.selected_lookback)
            self.plot_efficient_frontier(results)
            self.populate_weights_table(results)

            # Render correlation and covariance matrices (needed for export)
            _, daily_returns = self.compute_daily_returns(tickers, self.selected_lookback)
            corr = daily_returns.corr()
            self.populate_matrix_plot(self.corr_plot, self.corr_image, 'corr_text_items', corr, ".2f")
            cov = daily_returns.cov() * 252
            self.populate_matrix_plot(self.cov_plot, self.cov_image, 'cov_text_items', cov, ".4f")

            self.export_outputs(tickers)

            # Collapse left panel, show right toggle
            self.left_panel.hide()
            self.left_toggle.setText("▶")
            self.right_toggle.show()
        except Exception as e:
            QMessageBox.critical(
                self,
                "Calculation Error",
                f"Failed to calculate efficient frontier:\n{str(e)}",
            )
        finally:
            QApplication.restoreOverrideCursor()

    def populate_weights_table(self, results):
        """Fill the right-panel table with portfolio stats and weights."""
        tickers = results["tickers"]
        portfolios = [
            ("Max Sharpe", results["tangency_weights"], results["tangency_ret"],
             results["tangency_vol"], results["sharpe_ratio"]),
            # ("Max Sortino", results["sortino_weights"], results["sortino_ret"],
            #  results["sortino_vol"], results["sortino_ratio"]),
            ("Min Variance", results["min_vol_weights"], results["min_vol_ret"],
             results["min_vol_vol"], results["min_vol_vol"]),
        ]

        # Determine which tickers have weight >= 0.1% in any portfolio
        threshold = 0.001
        visible_tickers = set()
        for _, weights, _, _, _ in portfolios:
            for i, w in enumerate(weights):
                if w >= threshold:
                    visible_tickers.add(tickers[i])

        # Sort visible tickers by max weight across portfolios (descending)
        def max_weight(ticker):
            idx = tickers.index(ticker)
            return max(p[1][idx] for p in portfolios)
        visible_sorted = sorted(visible_tickers, key=max_weight, reverse=True)

        # Build rows: CAGR, Volatility, Ratio, separator, Weights header, then ticker weights
        summary_rows = ["CAGR", "StDev", "Ratio", "", ""]
        row_labels = summary_rows + visible_sorted
        num_rows = len(row_labels)

        self.weights_table.setRowCount(num_rows)
        self.weights_table.setVerticalHeaderLabels(row_labels)

        header_color = QColor(60, 60, 60)
        separator_color = QColor(40, 40, 40)

        for col, (name, weights, ret, vol, ratio) in enumerate(portfolios):
            # CAGR
            item = QTableWidgetItem(f"{ret:.2%}")
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setBackground(QBrush(header_color))
            self.weights_table.setItem(0, col, item)

            # Volatility
            item = QTableWidgetItem(f"{vol:.2%}")
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setBackground(QBrush(header_color))
            self.weights_table.setItem(1, col, item)

            # Ratio
            ratio_str = f"{ratio:.4f}" if name == "Min Variance" else f"{ratio:.2f}"
            item = QTableWidgetItem(ratio_str)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setBackground(QBrush(header_color))
            self.weights_table.setItem(2, col, item)

            # Separator row
            item = QTableWidgetItem("")
            item.setBackground(QBrush(separator_color))
            self.weights_table.setItem(3, col, item)

            # Ticker weights
            for row_offset, ticker in enumerate(visible_sorted):
                idx = tickers.index(ticker)
                w = weights[idx]
                text = f"{w:.1%}" if w >= threshold else "—"
                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.weights_table.setItem(5 + row_offset, col, item)

        # "Weights" header spanning all 3 columns
        self.weights_table.setSpan(4, 0, 1, 3)
        weights_header = QTableWidgetItem("Weights")
        weights_header.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        weights_header.setBackground(QBrush(header_color))
        font = weights_header.font()
        font.setBold(True)
        weights_header.setFont(font)
        self.weights_table.setItem(4, 0, weights_header)

        self.weights_table.resizeRowsToContents()
        self.weights_table.setRowHeight(3, 6)  # thin separator
        self.weights_table.show()

    def plot_efficient_frontier(self, results):
        self.plot_widget.clear()

        legend = self.plot_widget.addLegend(offset=(10, 10))
        legend.clear()

        # --- Simulated portfolios colored by Sharpe ratio ---
        sharpes = np.array(results["sim_sharpes"])
        s_min, s_max = sharpes.min(), sharpes.max()
        if s_max == s_min:
            normed = np.zeros_like(sharpes)
        else:
            normed = (sharpes - s_min) / (s_max - s_min)

        cmap = pg.colormap.get('magma')
        colors = cmap.map(normed, mode='byte')  # Nx4 uint8 array
        colors[:, 3] = 128  # 50% alpha

        brushes = [QBrush(QColor(int(c[0]), int(c[1]), int(c[2]), int(c[3]))) for c in colors]

        sim_scatter = pg.ScatterPlotItem(
            x=np.array(results["sim_vols"]),
            y=np.array(results["sim_rets"]),
            size=5,
            brush=brushes,
            pen=pg.mkPen(None),
        )
        self.plot_widget.addItem(sim_scatter)

        # --- Efficient frontier curve (skip for 2-asset case) ---
        if len(results["tickers"]) > 2:
            frontier_pen = pg.mkPen(color=(50, 205, 50), width=2)
            self.plot_widget.plot(
                x=np.array(results["frontier_vols"]),
                y=np.array(results["frontier_rets"]),
                pen=frontier_pen,
                name="Efficient Frontier",
            )

        # --- Individual asset labels (text only, no dots) ---
        label_font = QFont()
        label_font.setPointSize(10)
        label_font.setBold(True)
        for i, ticker in enumerate(results["tickers"]):
            text = pg.TextItem(text=ticker, color=(220, 220, 220), anchor=(0.5, 0.5))
            text.setFont(label_font)
            text.setPos(results["individual_vols"][i], results["individual_rets"][i])
            text.setZValue(10)
            self.plot_widget.addItem(text)

        # --- Tangency portfolio ---
        tangency_scatter = pg.ScatterPlotItem(
            x=[results["tangency_vol"]],
            y=[results["tangency_ret"]],
            size=20,
            symbol='star',
            brush=QBrush(QColor('orange')),
            pen=pg.mkPen('w', width=1),
        )
        self.plot_widget.addItem(tangency_scatter)
        legend.addItem(tangency_scatter, "Tangency Portfolio")

        # --- Capital Market Line ---
        rf = results["risk_free_rate"]
        cml_x = np.linspace(0, max(results["frontier_vols"]) * 1.2, 100)
        sharpe = (results["tangency_ret"] - rf) / results["tangency_vol"]
        cml_y = rf + sharpe * cml_x
        cml_pen = pg.mkPen(color=(180, 180, 180), width=1.5, style=Qt.PenStyle.DashLine)
        self.plot_widget.plot(
            x=cml_x,
            y=cml_y,
            pen=cml_pen,
            name=f"CML (Rf={rf:.2%})",
        )

        # --- Zoom to data with small padding ---
        all_vols = results["sim_vols"] + results["individual_vols"] + results["frontier_vols"]
        all_rets = results["sim_rets"] + results["individual_rets"] + results["frontier_rets"]
        x_min, x_max = min(all_vols), max(all_vols)
        y_min, y_max = min(all_rets), max(all_rets)
        x_pad = (x_max - x_min) * 0.05
        y_pad = (y_max - y_min) * 0.05
        self.plot_widget.setXRange(max(0, x_min - x_pad), x_max + x_pad, padding=0)
        self.plot_widget.setYRange(y_min - y_pad, y_max + y_pad, padding=0)


def run_gui():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
